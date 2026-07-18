"""Build ONE unified product catalog from the two raw sources.

Inputs (docs/raw/, immutable):
  - products_detail.json  : 13,754 crawled products (name / price / image / rating + shallow spec).
  - Spec_cate_gia.xlsx    : deep technical specs, 14 category sheets (~8,700 rows, keyed by sku;
                            no product name / image).

Output (docs/dataset/):
  - catalog/catalog.jsonl        : ONE record per product — the merge of both sources.
        * products found in products_detail keep full commercial info (name/price/image/…);
          deep specs from the xlsx are merged into their `specs`.
        * products that exist ONLY in the xlsx are added too, with name/image = null
          (technical-reference rows — filter on `name` when you need customer-facing products).
  - catalog/catalog.index.json   : summary stats (counts, coverage, quality flags).

The xlsx is read directly (the intermediate docs/dataset/specs/ folder is no longer used), so the
catalog is fully reproducible from docs/raw/ alone.
"""
import json
import re
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
MASTER_PATH = ROOT / "docs" / "raw" / "products_detail.json"
XLSX_PATH = ROOT / "docs" / "raw" / "Spec_cate_gia.xlsx"
OUT_CATALOG = ROOT / "docs" / "dataset" / "catalog"

# xlsx columns that are identifiers / commercial metadata, not spec attributes
SPEC_META = {
    "model_code", "sku", "productidweb", "category_code", "brand_id", "brand",
    "giá gốc", "giá khuyến mãi", "khuyến mãi quà",
}


# ---------- value normalizers ----------

def to_price(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(round(v)) or None
    digits = re.sub(r"[^\d]", "", str(v))
    return int(digits) if digits else None


def to_rating(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


def to_sold(v):
    if v is None or v == "":
        return None
    s = str(v).strip().lower()
    mult = 1
    if s.endswith("k"):
        mult, s = 1000, s[:-1]
    s = s.replace(",", ".")
    try:
        return int(round(float(s) * mult))
    except ValueError:
        digits = re.sub(r"[^\d]", "", str(v))
        return int(digits) if digits else None


def to_colors(v):
    if not v:
        return []
    return [p.strip() for p in re.split(r"[,;/]| - ", str(v)) if p.strip()]


def clean_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# ---------- unified record shape ----------
# Every record has the same keys; source-specific fields are null when absent.

def empty_record():
    return {
        "product_id": None, "sku": None, "model_code": None, "productcode": None,
        "name": None, "brand": None,
        "category": {"id": None, "name": None},
        "price": {"original": None, "sale": None, "currency": "VND"},
        "rating": None, "quantity_sold": None, "colors": [],
        "image_url": None, "url": None,
        "warranty": None, "accessories": None, "promotion": None,
        "online_sale_only": False, "crawled_at": None,
        "specs": {},
    }


def from_products_detail(p):
    r = empty_record()
    r.update({
        "product_id": clean_str(p.get("product_id")),
        "productcode": clean_str(p.get("productcode")),
        "name": clean_str(p.get("tên sản phẩm")),
        "brand": clean_str(p.get("brand")),
        "category": {"id": p.get("category_id"), "name": clean_str(p.get("category_name"))},
        "price": {"original": to_price(p.get("Giá gốc")),
                  "sale": to_price(p.get("Giá khuyến mãi")), "currency": "VND"},
        "rating": to_rating(p.get("rating_vote")),
        "quantity_sold": to_sold(p.get("quantity_sold")),
        "colors": to_colors(p.get("màu sắc")),
        "image_url": clean_str(p.get("url_image")),
        "url": clean_str(p.get("url")),
        "warranty": clean_str(p.get("chính sách bảo hành")),
        "accessories": clean_str(p.get("Phụ kiện đi kèm")),
        "promotion": clean_str(p.get("promotion")),
        "online_sale_only": bool(p.get("onlineSaleOnly")),
        "crawled_at": clean_str(p.get("time_crawler")),
        "specs": {k: v for k, v in (p.get("spec_product") or {}).items()
                  if v is not None and str(v).strip() != ""},
    })
    return r


def from_spec_row(row, category_name):
    r = empty_record()
    r.update({
        "product_id": clean_str(row.get("productidweb")),
        "sku": clean_str(row.get("sku")),
        "model_code": clean_str(row.get("model_code")),
        "brand": clean_str(row.get("brand")),
        "category": {"id": None, "name": category_name},
        "price": {"original": to_price(row.get("giá gốc")),
                  "sale": to_price(row.get("giá khuyến mãi")), "currency": "VND"},
        "promotion": clean_str(row.get("khuyến mãi quà")),
        "specs": spec_fields(row),
    })
    return r


def spec_fields(row):
    return {k: v for k, v in row.items()
            if k not in SPEC_META and v is not None and str(v).strip() != ""}


# ---------- xlsx ----------

def load_spec_sheets():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else f"col_{i}"
                   for i, h in enumerate(next(rows))]
        for row in rows:
            if all(v is None for v in row):
                continue
            yield sheet_name, dict(zip(headers, row))


# ---------- quality flags (index stats only, not stored on record) ----------

def quality_flags(r):
    flags = []
    if r["price"]["original"] is None and r["price"]["sale"] is None:
        flags.append("missing_price")
    if not r["name"]:
        flags.append("missing_name")
    if not r["specs"]:
        flags.append("no_spec")
    return flags


def write_jsonl(path, records):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for r in records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def main():
    master = json.loads(MASTER_PATH.read_text(encoding="utf-8"))
    records = [from_products_detail(p) for p in master]
    by_id = {r["product_id"]: r for r in records if r["product_id"]}
    n_master = len(records)

    enriched = set()
    fields_added = 0
    spec_only = {}   # keyed by sku to keep each variant once
    spec_rows = 0
    for sheet_name, row in load_spec_sheets():
        spec_rows += 1
        pid = clean_str(row.get("productidweb"))
        target = by_id.get(pid) if pid else None
        if target is not None:
            for k, v in spec_fields(row).items():
                if k not in target["specs"]:
                    target["specs"][k] = v
                    fields_added += 1
            if target["sku"] is None:
                target["sku"] = clean_str(row.get("sku"))
            if target["model_code"] is None:
                target["model_code"] = clean_str(row.get("model_code"))
            enriched.add(target["product_id"])
        else:
            key = clean_str(row.get("sku")) or f"{sheet_name}:{spec_rows}"
            spec_only[key] = from_spec_row(row, sheet_name)

    records.extend(spec_only.values())

    write_jsonl(OUT_CATALOG / "catalog.jsonl", records)

    cat_counts = Counter(r["category"]["name"] for r in records)
    brand_counts = Counter(r["brand"] for r in records if r["brand"])
    flag_counts = Counter(f for r in records for f in quality_flags(r))
    index = {
        "generated_from": ["docs/raw/products_detail.json", "docs/raw/Spec_cate_gia.xlsx"],
        "total_products": len(records),
        "from_products_detail": n_master,
        "spec_only_added": len(spec_only),
        "enriched_with_deep_specs": len(enriched),
        "deep_spec_fields_added": fields_added,
        "spec_source_rows": spec_rows,
        "categories": {"count": len(cat_counts), "top": cat_counts.most_common(25)},
        "brands": {"count": len(brand_counts), "top": brand_counts.most_common(25)},
        "flags": dict(flag_counts),
        "notes": (
            "One unified dataset. Records with name=null are technical-reference products that "
            "exist only in Spec_cate_gia.xlsx (no commercial info) — filter on `name` for "
            "customer-facing products."
        ),
    }
    (OUT_CATALOG / "catalog.index.json").write_text(
        json.dumps(index, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"catalog.jsonl        : {len(records)} products (unified)")
    print(f"  from products_detail : {n_master}")
    print(f"  spec-only added      : {len(spec_only)} (name=null, technical reference)")
    print(f"  enriched deep specs  : {len(enriched)} products (+{fields_added} fields)")
    print(f"  flags                : {dict(flag_counts)}")
    print(f"-> {OUT_CATALOG}/")


if __name__ == "__main__":
    main()
