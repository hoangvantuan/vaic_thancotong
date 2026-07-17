"""Convert Spec_cate_gia.xlsx (1 sheet per product category) into 1 JSON file per sheet."""
import json
import re
import unicodedata
from pathlib import Path

import openpyxl

SRC = Path(__file__).resolve().parent.parent / "docs" / "raw" / "Spec_cate_gia.xlsx"
OUT_DIR = Path(__file__).resolve().parent.parent / "docs" / "converted"


def slugify(name: str) -> str:
    name = name.replace("đ", "d").replace("Đ", "D")
    name = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    name = re.sub(r"[^a-zA-Z0-9]+", "_", name).strip("_").lower()
    return name


def sheet_to_records(ws) -> list[dict]:
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(next(rows))]
    records = []
    for row in rows:
        if all(v is None for v in row):
            continue
        records.append(dict(zip(headers, row)))
    return records


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        records = sheet_to_records(ws)
        out_path = OUT_DIR / f"{slugify(sheet_name)}.json"
        with out_path.open("w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        print(f"{sheet_name!r} -> {out_path.name} ({len(records)} records)")


if __name__ == "__main__":
    main()
